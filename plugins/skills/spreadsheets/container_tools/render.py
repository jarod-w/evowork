#!/usr/bin/env python3
"""内容 JSON → xlsx / csv。08 §5.2 的第二个技能。

## 这个技能的立身之本：**公式而非硬编码结果**

08 §5.2 把它列为关键质量点，而它的落点就在这个文件与 schema 里：
计算列在内容 JSON 里给的是**公式模板**（`"=B{row}*C{row}"`），渲染器按行展开写进单元格。

为什么这条比"数字对不对"更重要：用户拿到表之后会改数。如果单元格里是算好的常量，
改了输入列，合计与占比纹丝不动 —— 而它看起来完全正常。这种错会一路带进汇报。
所以 schema 强制"给了 formula 的列，rows 里对应位置必须是 null"，
渲染器再校验一次（schema 表达不了这种跨字段约束）。

csv 属于**基础包**（不需要办公扩展），xlsx 需要 openpyxl。这一点在 SKILL.md 里说清了，
因为它决定了"没装扩展时还能不能干活"。
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

SKILL_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(SKILL_ROOT.parent / "_shared"))

from evowork_skill import (  # noqa: E402
    ensure_office_runtime,
    EXIT_INVALID_CONTENT,
    EXIT_RUNTIME_MISSING,
    fail,
    read_content,
    runtime_missing_message,
    succeeded,
    validate_content,
)

MAX_COLUMNS = 60


def load_template() -> dict:
    return read_content(SKILL_ROOT / "templates" / "default" / "template.json")


def column_letter(index: int) -> str:
    """0 → A、25 → Z、26 → AA。自己写是为了让 csv 路径不依赖 openpyxl。"""
    letters = ""
    index += 1
    while index > 0:
        index, rem = divmod(index - 1, 26)
        letters = chr(ord("A") + rem) + letters
    return letters


def validate(content: dict) -> None:
    validate_content(content, SKILL_ROOT / "schema" / "content.schema.json")

    for sheet_index, sheet in enumerate(content["sheets"]):
        columns = sheet["columns"]
        where = f"sheets/{sheet_index}"
        formula_positions = [i for i, c in enumerate(columns) if c.get("formula")]

        for row_index, row in enumerate(sheet["rows"]):
            if len(row) != len(columns):
                fail(
                    EXIT_INVALID_CONTENT,
                    f"{where}/rows/{row_index}: 这一行有 {len(row)} 个单元格，"
                    f"但 columns 有 {len(columns)} 列，两者必须一样长"
                    f"（计算列的位置写 null）。",
                )
            for position in formula_positions:
                if row[position] is not None:
                    # 这条就是这个技能存在的理由，所以报错要说清"为什么"
                    fail(
                        EXIT_INVALID_CONTENT,
                        f"{where}/rows/{row_index}/{position}: 「{columns[position]['header']}」"
                        f"是计算列（有 formula），这里必须是 null。"
                        f"填算好的数会让用户改了输入列之后结果不更新，而且看不出来。",
                    )

        headers = [c["header"] for c in columns]
        if len(set(headers)) != len(headers):
            fail(EXIT_INVALID_CONTENT, f"{where}/columns: 表头有重名，公式与条件格式会指错列。")

        for fmt_index, fmt in enumerate(sheet.get("conditional_formats", [])):
            if fmt["column"] not in headers:
                fail(
                    EXIT_INVALID_CONTENT,
                    f"{where}/conditional_formats/{fmt_index}: 找不到列「{fmt['column']}」。",
                )


def expand_formula(template: str, row_number: int) -> str:
    return template.replace("{row}", str(row_number))


def write_csv(sheet: dict, out_path: Path) -> None:
    """csv 没有公式，所以计算列**留空并在表头标注**，而不是悄悄算一个值填进去。

    这是一次显式降级（D2 的同一条原则）：用户看到空列会去问，看到一个算好的数不会。
    """
    columns = sheet["columns"]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8-sig", newline="") as handle:
        # utf-8-sig：Excel 打开无 BOM 的 UTF-8 csv 会把中文显示成乱码
        writer = csv.writer(handle)
        writer.writerow(
            [f"{c['header']}（公式列，csv 不支持）" if c.get("formula") else c["header"] for c in columns]
        )
        for row in sheet["rows"]:
            writer.writerow(["" if value is None else value for value in row])


def render_xlsx(content: dict, out_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import CellIsRule, DataBarRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        fail(EXIT_RUNTIME_MISSING, runtime_missing_message("office", "生成 xlsx"))

    template = load_template()
    formats = template["number_formats"]
    widths = template["default_widths"]
    header_spec = template["header"]
    colors = template["colors"]

    book = Workbook()
    book.remove(book.active)

    for sheet_spec in content["sheets"]:
        sheet = book.create_sheet(sheet_spec["name"])
        columns = sheet_spec["columns"]

        for index, column in enumerate(columns):
            cell = sheet.cell(row=1, column=index + 1, value=column["header"])
            cell.font = Font(
                bold=header_spec["bold"],
                color=header_spec["font_color"],
                name=header_spec["font_cjk"],
            )
            cell.fill = PatternFill("solid", fgColor=header_spec["fill"])
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            kind = column.get("type", "text")
            sheet.column_dimensions[get_column_letter(index + 1)].width = column.get(
                "width", widths[kind]
            )

        for row_index, row in enumerate(sheet_spec["rows"]):
            excel_row = row_index + 2
            for col_index, column in enumerate(columns):
                kind = column.get("type", "text")
                if column.get("formula"):
                    value = expand_formula(column["formula"], excel_row)
                else:
                    value = row[col_index]
                cell = sheet.cell(row=excel_row, column=col_index + 1, value=value)
                cell.number_format = formats[kind]
                if kind != "text":
                    cell.font = Font(name=header_spec["font_cjk"])

        last_row = len(sheet_spec["rows"]) + 1

        if sheet_spec.get("total_row") and last_row >= 2:
            total_row = last_row + 1
            sheet.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
            for col_index, column in enumerate(columns):
                if column.get("type", "text") in ("number", "integer", "currency"):
                    letter = get_column_letter(col_index + 1)
                    # 合计也是公式：用户删掉几行之后它要跟着变
                    cell = sheet.cell(
                        row=total_row,
                        column=col_index + 1,
                        value=f"=SUM({letter}2:{letter}{last_row})",
                    )
                    cell.number_format = formats[column.get("type", "number")]
                    cell.font = Font(bold=True, name=header_spec["font_cjk"])

        # 冻结首行默认开：表格超过一屏时不冻结就没法看
        if sheet_spec.get("freeze_header", True):
            sheet.freeze_panes = "A2"

        headers = [c["header"] for c in columns]
        for fmt in sheet_spec.get("conditional_formats", []):
            letter = get_column_letter(headers.index(fmt["column"]) + 1)
            span = f"{letter}2:{letter}{last_row}"
            rule = fmt["rule"]
            if rule == "negative-red":
                sheet.conditional_formatting.add(
                    span,
                    CellIsRule(operator="lessThan", formula=["0"], font=Font(color=colors["negative"])),
                )
            elif rule == "above-average-green":
                sheet.conditional_formatting.add(
                    span,
                    CellIsRule(
                        operator="greaterThan",
                        formula=[f"AVERAGE({span})"],
                        font=Font(color=colors["positive"]),
                    ),
                )
            elif rule == "top10-green":
                sheet.conditional_formatting.add(
                    span,
                    CellIsRule(
                        operator="greaterThanOrEqual",
                        formula=[f"LARGE({span},10)"],
                        font=Font(color=colors["positive"]),
                    ),
                )
            elif rule == "data-bar":
                sheet.conditional_formatting.add(span, DataBarRule(color=colors["data_bar"]))

        if sheet_spec.get("note"):
            sheet.cell(row=last_row + 3, column=1, value=sheet_spec["note"])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    book.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="内容 JSON → xlsx / csv")
    parser.add_argument("--content", required=True)
    parser.add_argument("--out", required=True, help="输出路径，扩展名决定格式（.xlsx / .csv）")
    parser.add_argument("--validate-only", action="store_true")
    args = parser.parse_args()

    # 缺办公扩展的模块时换到扩展的解释器重跑（见 evowork_skill.ensure_office_runtime）。
    # --validate-only 不需要它，但提前换掉更简单，也让两条路径的行为一致
    ensure_office_runtime(("openpyxl",))

    content = read_content(Path(args.content).resolve())
    validate(content)

    if args.validate_only:
        succeeded(sheets=len(content["sheets"]), validated=True)
        return

    out_path = Path(args.out).resolve()
    suffix = out_path.suffix.lower()
    if suffix == ".csv":
        if len(content["sheets"]) > 1:
            fail(
                EXIT_INVALID_CONTENT,
                "csv 只能装一张表，但内容里有多张。要么输出 xlsx，要么每张表各写一个 csv。",
            )
        write_csv(content["sheets"][0], out_path)
        succeeded(out_path, rows=len(content["sheets"][0]["rows"]), format="csv")
        return
    if suffix != ".xlsx":
        fail(EXIT_INVALID_CONTENT, f"输出格式只支持 .xlsx 与 .csv，收到 {suffix or '(无扩展名)'}。")

    render_xlsx(content, out_path)
    succeeded(out_path, sheets=len(content["sheets"]), format="xlsx")


if __name__ == "__main__":
    main()
